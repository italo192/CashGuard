import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar



# Função para salvar o valor
def salvar_valor():
    valor_dia = float(entry_valor.get())
    valores.append(valor_dia)
    total_valores = sum(valores)
    entry_valor.delete(0, tk.END)
    label_status.config(text="Valor salvo com sucesso!", foreground="green")
    label_total.config(text=f"Total economizado: R${total_valores:.2f}")

    # Adiciona valor na planilha
    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1, value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha, column=2, value=valor_dia)

    # Salva a planilha
    workbook.save("valores_diarios.xlsx")

# Função para plotar os gráficos
def plotar_grafico():
    # Verifica se a planilha contém pelo menos uma linha de dados
    if sheet.max_row <= 1:
        print(
            "A planilha está vazia ou contém apenas o cabeçalho. Não é possível gerar gráficos."
        )
        return

    # Obtem as datas e valores
    datas = [
        cell.value.date()
        if isinstance(cell.value, datetime)
        else datetime.strptime(cell.value, "%d-%m-%y").date()
        for cell in sheet["A"][1:]
    ]
    valores = [cell.value for cell in sheet["B"][1:]]

    # Calcular maior e menor valor da planilha se houver dados
    if valores:
        maior_valor = max(valores)
        menor_valor = min(valores)
    else:
        maior_valor = menor_valor = 0

    # Agrupa valores por mês
    dados_mensais = {}
    for data, valor in zip(datas, valores):
        mes_ano = data.strftime("%m-%Y")
        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]

    # Plotar o gráfico de barras e pizza em um único Figure
    fig, (ax_barras, ax_pizza) = plt.subplots(
        1, 2, figsize=(6, 2.5), gridspec_kw={"width_ratios": [3, 2]}
    )
    fig.patch.set_facecolor("#ffffff")  # Cor de fundo do Figure

    # Gráfico de barras
    barras = ax_barras.bar(
        range(len(dados_mensais)), [sum(valores) for valores in dados_mensais.values()]
    )

    for i, barra in enumerate(barras):
        altura = barra.get_height()
        ax_barras.text(
            barra.get_x() + barra.get_width() / 2,
            altura,
            f"R${altura:.2f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )

    nomes_meses = []
    for mes_ano in dados_mensais.keys():
        mes, ano = mes_ano.split("-")
        nome_mes = calendar.month_name[int(mes)]
        nomes_meses.append(f"{nome_mes}-{ano}")

    ax_barras.set_xticks(range(len(dados_mensais)))
    ax_barras.set_xticklabels(nomes_meses, ha="right", fontsize=8, rotation=45)

    ax_barras.spines["top"].set_visible(False)
    ax_barras.spines["right"].set_visible(False)
    ax_barras.spines["bottom"].set_visible(False)
    ax_barras.spines["left"].set_visible(False)

    ax_barras.set_title("Economia por Mês", fontsize=10)
    ax_barras.title.set_position([0.5, 1.05])
    ax_barras.set_xlabel("Mês", fontsize=8)
    ax_barras.set_ylabel("Valor Economizado", fontsize=8)

    # Adicionar valores e datas ao lado do gráfico de barras
    text_vals = "\n".join(
        [f"{data}: R${valor:.2f}" for data, valor in zip(datas, valores)]
    )
    label_valores.config(text=text_vals, font=("Arial", 8))

    # Gráfico de pizza
    # Calcular maior e menor valor da planilha
    maior_valor = max(valores)
    menor_valor = min(valores)

    # Calcular porcentagens
    porcentagem_maior = (maior_valor / total_valores) * 100
    porcentagem_menor = (menor_valor / total_valores) * 100

    # Ajuste da organização do gráfico de pizza
    wedges, texts, autotexts = ax_pizza.pie(
        [maior_valor, menor_valor],
        labels=None,
        autopct="",
        startangle=90,
        wedgeprops=dict(width=0.3),
    )

    # Ajustes visuais do gráfico de pizza
    ax_pizza.legend(
        wedges,
        [
            f"Maior Valor\n{porcentagem_maior:.1f}% do total",
            f"Menor Valor\n{porcentagem_menor:.1f}% do total",
        ],
        loc="center left",
        bbox_to_anchor=(0.9, 0, 0.5, 1),
        fontsize=8,
    )
    ax_pizza.axis("equal")  # Garante que o gráfico de pizza seja um círculo

    # Ajustar o layout do Figure
    plt.tight_layout()

    # Exibir o Figure em um único Canvas
    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.get_tk_widget().grid(
        row=6, column=0, columnspan=2, padx=10, pady=10, sticky="nsew"
    )
    canvas.draw()


# GUI
janela = tk.Tk()
janela.title("App de Poupança Pessoal")
janela.geometry("700x500")
janela.configure(bg="#252525")

style = ttk.Style()
style.theme_use("clam")
style.configure(
    "TLabel", background="#252525", foreground="#FFFFFF", font=("Arial", 12)
)
style.configure("TEntry", fieldbackground="#FFFFFF", font=("Arial", 12))
style.configure(
    "TButton", background="#4CAF50", foreground="#FFFFFF", font=("Arial", 12)
)

label_instrucao = ttk.Label(janela, text="Insira um Valor")
label_status = ttk.Label(janela, text="Valor de Hoje", foreground="red")
label_total = ttk.Label(janela, font=("Arial", 14, "bold"))
entry_valor = ttk.Entry(janela)

button_salvar = ttk.Button(janela, text="Salvar", command=salvar_valor)
label_valores = ttk.Label(janela, text="", font=("Arial", 8))

# posicionamento dos elementos
label_instrucao.grid(row=0, column=0, pady=10)
entry_valor.grid(row=1, column=0, pady=5)
button_salvar.grid(row=2, column=0, pady=10)
label_status.grid(row=3, column=0)
label_total.grid(row=4, column=0, pady=10)
button_salvar.grid(row=5, column=0, pady=10)

# Carregamento da planilha
try:
    workbook = load_workbook("planilha.xlsx")
except FileNotFoundError:
    workbook = Workbook()

# Primeira planilha
sheet = workbook.active

# Verificar se a planilha já possui valores salvos
if sheet.max_row == 0:
    sheet.cell(row=1, column=1, value="Data")
    sheet.cell(row=1, column=2, value="Valor diário")

# Valores salvos na lista
valores = [cell.value for cell in sheet["B"][1:]]

# Total economizado
total_valores = sum(valores)
label_total.config(text=f"Total economizado: R${total_valores:.2f}")

# Configurar a expansão da célula
for i in range(7):
    janela.grid_rowconfigure(i, weight=1)
    janela.grid_columnconfigure(i, weight=1)


# Chama a função plotar_grafico() apenas se a planilha contiver dados
if sheet.max_row > 0:
    plotar_grafico()

janela.mainloop()
